# myapp/admin.py
from django.contrib import admin
from django.utils.html import format_html
from django.http import HttpResponse
from django.urls import path, reverse
from django.shortcuts import get_object_or_404, render
from django.utils.safestring import mark_safe
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime

from .models import Specialist, CourseCategory, Course
from modeltranslation.admin import TranslationAdmin

@admin.register(CourseCategory)
class CourseCategoryAdmin(TranslationAdmin):
    list_display = ('title',)

@admin.register(Course)
class CourseAdmin(TranslationAdmin):
    list_display = ('title', 'category')
    list_filter = ('category',)
    search_fields = ('title', 'short_description')

@admin.register(Specialist)
class SpecialistAdmin(TranslationAdmin):
    list_display = (
        'full_name', 'city', 'specialization', 'position', 
        'is_member', 'moderation_status', 'created_at', 'photo_preview', 'detail_link'
    )
    search_fields = ('full_name', 'email', 'phone', 'specialization', 'workplace')
    list_filter = ('city', 'is_member', 'category', 'moderation_status', 'created_at')
    readonly_fields = ('created_at', 'updated_at')
    
    fieldsets = (
        ('Личная информация', {
            'fields': ('full_name', 'email', 'phone', 'additional_phone', 'birth_date', 
                      'citizenship', 'city', 'social_links', 'photo')
        }),
        ('Образование', {
            'fields': ('university', 'faculty', 'study_years', 'degree')
        }),
        ('Профессиональная деятельность', {
            'fields': ('position', 'workplace', 'job_region', 'job_description', 
                      'experience_years', 'professional_description', 'specialization')
        }),
        ('Достижения', {
            'fields': ('skills', 'awards', 'publications')
        }),
        ('Мотивация и интересы', {
            'fields': ('motivation', 'interests', 'recommender')
        }),
        ('Файлы', {
            'fields': ('resume', 'documents')
        }),
        ('Дополнительная информация', {
            'fields': ('about', 'cases', 'category')
        }),
        ('Согласия', {
            'fields': ('confirm_data', 'consent_personal_data')
        }),
        ('Статус', {
            'fields': ('is_member', 'moderation_status', 'created_at', 'updated_at')
        })
    )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export-excel/', self.admin_site.admin_view(self.export_excel), name='specialist_export_excel'),
            path('<int:object_id>/detail/', self.admin_site.admin_view(self.detail_view), name='specialist_detail'),
        ]
        return custom_urls + urls

    def photo_preview(self, obj):
        if obj.photo:
            return format_html(
                '<img src="{}" style="width: 40px; height: 40px; object-fit: cover; border-radius: 4px;" />',
                obj.photo.url
            )
        return "-"
    photo_preview.short_description = "Фото"

    def detail_link(self, obj):
        url = reverse('admin:specialist_detail', args=[obj.pk])
        return format_html('<a href="{}" class="button">Подробнее</a>', url)
    detail_link.short_description = "Действия"

    def detail_view(self, request, object_id):
        specialist = get_object_or_404(Specialist, pk=object_id)
        
        # Парсим интересы из строки обратно в список
        interests_list = []
        if specialist.interests:
            interests_list = specialist.interests.split(', ')
        
        context = {
            'specialist': specialist,
            'interests_list': interests_list,
            'title': f'Детальная информация: {specialist.full_name}',
            'opts': self.model._meta,
        }
        return render(request, 'admin/specialist_detail.html', context)

    def export_excel(self, request):
        # Создаем Excel файл
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Заявки специалистов"

        # Заголовки
        headers = [
            'ФИО', 'Email', 'Телефон', 'Доп. телефон', 'Дата рождения', 'Гражданство',
            'Город', 'Соцсети', 'Учебное заведение', 'Факультет', 'Годы обучения',
            'Академическая степень', 'Должность', 'Место работы', 'Регион работы',
            'Описание компании', 'Стаж (лет)', 'Описание деятельности', 'Навыки',
            'Награды', 'Публикации', 'Мотивация', 'Интересы', 'Рекомендатель',
            'Специализация', 'Категория', 'Участник ассоциации', 'Статус модерации',
            'Дата подачи заявки', 'Подтверждение данных', 'Согласие на обработку'
        ]

        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Получаем данные
        specialists = Specialist.objects.all().order_by('-created_at')

        # Записываем данные
        for row, specialist in enumerate(specialists, 2):
            data = [
                specialist.full_name,
                specialist.email,
                specialist.phone,
                specialist.additional_phone,
                specialist.birth_date.strftime('%d.%m.%Y') if specialist.birth_date else '',
                specialist.citizenship,
                specialist.city,
                specialist.social_links,
                specialist.university,
                specialist.faculty,
                specialist.study_years,
                specialist.degree,
                specialist.position,
                specialist.workplace,
                specialist.job_region,
                specialist.job_description,
                specialist.experience_years,
                specialist.professional_description,
                specialist.skills,
                specialist.awards,
                specialist.publications,
                specialist.motivation,
                specialist.interests,
                specialist.recommender,
                specialist.specialization,
                specialist.category,
                'Да' if specialist.is_member else 'Нет',
                specialist.get_moderation_status_display(),
                specialist.created_at.strftime('%d.%m.%Y %H:%M'),
                'Да' if specialist.confirm_data else 'Нет',
                'Да' if specialist.consent_personal_data else 'Нет',
            ]
            
            for col, value in enumerate(data, 1):
                worksheet.cell(row=row, column=col, value=value)

        # Автоширина колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Создаем HTTP ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"specialists_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        workbook.save(response)
        return response

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['export_url'] = reverse('admin:specialist_export_excel')
        return super().changelist_view(request, extra_context)